"""Collection of various functions used to analyze data."""
import logging
import os
import pickle
import re
import time

import requests

from collections import Counter
from collections import OrderedDict
from collections import defaultdict

from openpyxl import Workbook
from openpyxl import load_workbook

from follower_analyzer.db import STATUS_COLUMNS
from follower_analyzer import db2
from follower_analyzer.db2 import USER_COLUMNS2


OBAMA = 'BarackObama'
TRUMP = 'realDonaldTrump'
BOTH = 'BOTH'
ALL = 'ALL'
STRING_SUFFIXED_WITH_DIGITS_REGEX = re.compile(r'(.+?)(\d+)$')


def _save_data(fpath: str, obj: object):
    """Save an arbitrary Python object. Usually a result of some analysis."""
    with open(fpath, 'wb') as pfd:
        pickle.dump(obj, pfd)


def _init_counters():
    return OrderedDict({OBAMA: defaultdict(Counter), BOTH: defaultdict(Counter), TRUMP: defaultdict(Counter)})


def collect_counters(dbpath: str, max_iterations: int=1000, bucket: int=10, query_limit: int=100000):
    """To get an easy high-level overview of the collected data per DB. Basically does a table scan, and depending on
    whether a column is an ENUM or a count that is or isn't too granular to require bucketing counts them in different
    ways. The structure of the result is
    OrderedDict({OBAMA/BOTH/TRUMP: {column_name: Counter({ENUM/BUCKET/NUM: count})}})"""
    counters = _init_counters()
    start = time.time()
    statuses = {BOTH: set(), TRUMP: set()}
    conn = db2.get_conn(dbpath)
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM users')
    index = 0
    while True:
        if index == max_iterations:
            break
        else:
            index += 1
        rows = cursor.fetchmany(query_limit)
        for row in rows:
            following = row['following']
            if following == BOTH:
                names = [OBAMA, TRUMP, BOTH]
            else:
                names = [following]

            if following != OBAMA:
                if row['status'] != '':
                    statuses[following].add(row['status'])
                if row['previous_status'] != '':
                    statuses[following].add(row['previous_status'])

            for name in names:
                for k, v in USER_COLUMNS2.items():
                    if k in ['location', 'following']:
                        # TODO: Some of these need special handling.
                        pass
                    elif v == bool or k in ['utc_offset', 'lang'] or k.endswith('_color') or \
                            k.startswith('withheld_'):
                        counters[name][k][row[k]] += 1
                    elif v == int:
                        if row[k] < 10:
                            counters[name][k][row[k]] += 1
                        else:
                            counters[name][k][(row[k] // bucket) * bucket] += 1
                    elif k == 'created_at':
                        date = time.strftime('%Y%m%d', time.strptime(row[k], '%a %b %d %H:%M:%S +0000 %Y'))
                        counters[name][k][date] += 1
                    elif k in ['status', 'previous_status']:
                        if row[k] != '':
                            counters[name][k][1] += 1
                        else:
                            counters[name][k][0] += 1

                followers = row['followers_count']
                friends = row['friends_count']
                if followers > 0 and len(counters[name]['friends_to_followers']) < 10000:
                    ratio = friends / followers
                    if ratio < 1:
                        counters[name]['friends_to_followers'][round(ratio, 1)] += 1
                    else:
                        counters[name]['friends_to_followers'][round(ratio)] += 1
                if friends > 0 and len(counters[name]['followers_to_friends']) < 10000:
                    ratio = followers / friends
                    if ratio < 1:
                        counters[name]['followers_to_friends'][round(ratio, 1)] += 1
                    else:
                        counters[name]['followers_to_friends'][round(ratio)] += 1

        if len(rows) < query_limit:
            break

    cursor.execute('SELECT * FROM statuses')
    index = 0
    while True:
        if index == max_iterations:
            break
        else:
            index += 1
        rows = cursor.fetchmany(query_limit)
        for row in rows:
            if row['id_str'] in statuses[BOTH]:
                names = [OBAMA, TRUMP, BOTH]
            elif row['id_str'] in statuses[TRUMP]:
                names = [TRUMP]
            else:
                names = [OBAMA]

            for name in names:
                for k, v in STATUS_COLUMNS.items():
                    if k in ['contributors', 'lang', 'possibly_sensitive', 'scopes', 'withheld_copyright',
                             'withheld_in_countries', 'withheld_scope']:
                        counters[name]['status_{}'.format(k)][row[k]] += 1
                    elif k == 'created_at':
                        date = time.strftime('%Y%m%d', time.strptime(row[k], '%a %b %d %H:%M:%S +0000 %Y'))
                        counters[name]['status_{}'.format(k)][date] += 1

        if len(rows) < query_limit:
            break

    _save_data(dbpath.replace('.db', '.dat'), counters)
    logging.debug('Collecting various counts of users and statuses in %s took %s seconds.',
                  os.path.basename(dbpath), time.time() - start)


def _load_data(fpath: str):
    """Load an arbitrary Python object previously saved using _save_data."""
    with open(fpath, 'rb') as pfd:
        return pickle.load(pfd)


def _finalize_counters(counters: OrderedDict):
    """In addition to the OBAMA, BOTH, TRUMP counters, also derives an ALL from them."""
    if ALL in counters:
        return
    counters_for_all = defaultdict(Counter)
    for name, breakdown in counters.items():
        if name is 'BOTH':
            factor = -1
        else:
            factor = 1
        for column, column_counter in breakdown.items():
            for k, v in column_counter.items():
                counters_for_all[column][k] += (v * factor)
    counters[ALL] = counters_for_all


def merge_counters(counter_paths: list, merged_counter_path: str):
    """Merges all the counters saved by collect_counters to get an aggregatve view across all twitter_YYYYMM.dbs."""
    counters = _init_counters()
    for cpath in counter_paths:
        start = time.time()
        for name, breakdown in _load_data(cpath).items():
            for column, column_counter in breakdown.items():
                for k, v in column_counter.items():
                    counters[name][column][k] += v
        logging.debug('Merging various counts of users and statuses in %s took %s seconds.',
                      os.path.basename(cpath), time.time() - start)

    _finalize_counters(counters)
    _save_data(merged_counter_path, counters)


def _get_key_range(counters: OrderedDict, key: str):
    if key.endswith('_color') or key.endswith('lang'):
        return [mc[0] for mc in counters[ALL][key].most_common(10000)]
    else:
        key_range = set()
        for name in counters:
            key_range |= set(counters[name][key].keys())
        if '' in key_range:
            key_range.remove('')
            return [''] + sorted(key_range)
        else:
            return sorted(key_range)


def _get_title(key: str):
    return key.replace('_', ' ').title()


def _set_column(sheet, column: int, values: list):
    for index, val in enumerate(values):
        sheet.cell(row=index+1, column=column).value = val


def _set_row(sheet, row: int, values: tuple):
    for index, val in enumerate(values):
        sheet.cell(row=row, column=index+1).value = val


def _fill_sheet(sheet, counters: OrderedDict, keys: list):
    index = 1
    for key in keys:
        key_range = _get_key_range(counters, key)
        _set_column(sheet, index, [_get_title(key)] + key_range)
        for nindex, name in enumerate(counters):
            _set_column(sheet, index+1+nindex, [name] + [counters[name][key][k] for k in key_range])
        index += 6


def write_report(merged_counter_path: str, report_path: str):
    """Writes a huge XLSX with multiple sheets, filled using the result of merge_counters above."""
    counters = _load_data(merged_counter_path)

    workbook = Workbook()

    counts_sheet = workbook.active
    counts_sheet.title = 'Counts'
    _fill_sheet(counts_sheet, counters, ['favourites_count', 'followers_count', 'friends_count', 'listed_count',
                                         'statuses_count'])

    colors_sheet = workbook.create_sheet('Colors')
    _fill_sheet(colors_sheet, counters, ['profile_background_color', 'profile_link_color', 'profile_sidebar_fill_color',
                                         'profile_text_color'])

    time_sheet = workbook.create_sheet('Times')
    _fill_sheet(time_sheet, counters, ['created_at', 'status_created_at', 'utc_offset'])

    lang_sheet = workbook.create_sheet('Languages')
    _fill_sheet(lang_sheet, counters, ['lang', 'status_lang'])

    ratios_sheet = workbook.create_sheet('Ratios')
    _fill_sheet(ratios_sheet, counters, ['followers_to_friends', 'friends_to_followers'])

    flags_sheet = workbook.create_sheet('Flags & Miscellaneous')
    _set_row(flags_sheet, 1, ('Flag', 'Value') + tuple(counters.keys()))
    index = 2
    for key in ['contributors_enabled', 'default_profile', 'default_profile_image', 'geo_enabled',
                'profile_background_tile', 'protected', 'verified', 'status', 'previous_status',
                'status_possibly_sensitive', 'status_contributors', 'status_scopes', 'status_withheld_copyright',
                'status_withheld_in_countries', 'status_withheld_scope', 'withheld_in_countries', 'withheld_scope']:
        for val in _get_key_range(counters, key):
            _set_row(flags_sheet, index, (_get_title(key), val) + tuple(counters[name][key][val] for name in counters))
            index += 1

    workbook.save(report_path)


def create_status_following_index(dbpath: str, index_path: str):
    """Corresponding to a twitter_YYYYMM.db, creates a twitter_YYYYMM.index which is in the format
    {BOTH: set(status_ids_of_users_following_both), TRUMP: set(status_ids_of_users_following_trump)}"""
    start = time.time()
    conn = db2.get_conn(dbpath)
    cursor = conn.cursor()
    statuses = {BOTH: set(), TRUMP: set()}
    for name in [BOTH, TRUMP]:
        for column in ['status', 'previous_status']:
            cursor.execute("SELECT {} FROM users WHERE {} != '' AND following = '{}'".format(column, column, name))
            statuses[name] |= {row[0] for row in cursor.fetchall()}
    _save_data(index_path, statuses)
    logging.debug('Creating index of status to following (Trump/Both) for %s took %s seconds.',
                  os.path.basename(dbpath), time.time() - start)


def get_status_sources(dbpath: str, status_following_index_path: str, sources: dict):
    """Returns a {source: {OBAMA: count, TRUMP: count, BOTH: count, ALL: count}}. The source list comes from
    the statuses table, and the following comes from the status_following_index created above. This is because
    of the one way linking between users and statuses tables: users.{status, previous_status} -> statuses.id_str"""
    start = time.time()
    statuses = _load_data(status_following_index_path)
    with db2.get_conn(dbpath) as conn:
        cursor = conn.cursor()
        cursor.execute('SELECT id_str, source FROM statuses')
        for row in cursor.fetchall():
            source = row['source']
            if source not in sources:
                sources[source] = OrderedDict({OBAMA: 0, TRUMP: 0, BOTH: 0, ALL: 0})
            sources[source][ALL] += 1
            if row['id_str'] in statuses[BOTH]:
                sources[source][OBAMA] += 1
                sources[source][TRUMP] += 1
                sources[source][BOTH] += 1
            elif row['id_str'] in statuses[TRUMP]:
                sources[source][TRUMP] += 1
            else:
                sources[source][OBAMA] += 1
    logging.debug('Getting status sources from %s took %s seconds.',
                  os.path.basename(dbpath), time.time() - start)
    return sources


def unshorten_url(url: str) -> [str, None]:
    """Unshortens a URL following https://stackoverflow.com/a/28918160. Also strips the URL of prefixes like
    http://, https://, www., and suffix '/'."""
    session = requests.Session()
    try:
        return session.head(url, allow_redirects=True, timeout=5).url.replace('http://', '')\
                      .replace('https://', '').replace('www.', '').rstrip('/')
    except requests.exceptions.Timeout:
        return 'Timeout'
    except requests.exceptions.TooManyRedirects:
        return 'TooManyRedirects'
    except requests.exceptions.ConnectionError:
        return 'ConnectionError'
    except:
        return None


def add_sources_sheet(sources_path: str, report_path: str):
    """Writes the result of get_status_sources which was subsequently extended with urls from unshorten_url, as
    a new sheet to the XLSX created by write_report."""
    workbook = load_workbook(report_path)
    srcs_sheet = workbook.create_sheet('Sources')
    _set_row(srcs_sheet, 1, ('Source', 'URL', OBAMA, TRUMP, BOTH, ALL, 'UNSHORTENED'))
    index = 2
    for full_source_str, source_counter in sorted(_load_data(sources_path).items(), key=lambda x: x[1][ALL],
                                                  reverse=True):
        if full_source_str == '':
            source = url = ''
        else:
            url, source = full_source_str.replace('<a href="', '').replace('" rel="nofollow"', '')\
                                         .replace('</a>', '').split('>')
        try:
            _set_row(srcs_sheet, index, (source, url, source_counter[OBAMA], source_counter[TRUMP],
                                         source_counter[BOTH], source_counter[ALL], source_counter.get('UNSHORTENED')))
        except:
            if url == 'http://tweetcaster.com':
                _set_row(srcs_sheet, index, ('TweetCaster^Lfor Android', url, source_counter[OBAMA],
                                             source_counter[TRUMP], source_counter[BOTH], source_counter[ALL],
                                             source_counter.get('UNSHORTENED', '')))
            else:
                logging.error('Failed to write row. %s, %s', full_source_str, source_counter)
                raise
        index += 1
    workbook.save(report_path)


def create_screen_name_index(dbpath: str, screen_name_index_path: str):
    """NOTE: This function is treating the names as disjoint collections, the way stored in DB, to save on space."""
    start = time.time()
    with db2.get_conn(dbpath) as conn:
        cursor = conn.cursor()
        screen_names = {}
        for name in [OBAMA, TRUMP, BOTH]:
            cursor.execute('SELECT screen_name FROM users WHERE following = "{}"'.format(name))
            screen_names[name] = [row[0] for row in cursor.fetchall()]
        _save_data(screen_name_index_path, screen_names)
    logging.debug('Creating screen name index for %s took %s seconds',
                  os.path.basename(dbpath), time.time() - start)


def count_screen_names_ending_in_digits(screen_name_index_path: str, screen_names: dict):
    for name, value in _load_data(screen_name_index_path).items():
        for screen_name in value:
            match = re.match(STRING_SUFFIXED_WITH_DIGITS_REGEX, screen_name)
            if match is not None:
                screen_names[name][match.groups()[0].lower()] += 1
